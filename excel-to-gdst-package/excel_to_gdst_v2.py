# -*- coding: utf-8 -*-
"""
从 Excel 文件生成 GDST 文件 - 动态表头版本
用法: python excel_to_gdst_v2.py [excel文件]

Excel 格式要求:
- 第一行可以是元数据（如 F 描述 f:CommonRuleFact f）
- 第二行是表头行（包含：规则名、条件、其他条件列...、结果）
- 表头中不含"描述"列时，从第一行元数据行的第二列获取
- 每个 sheet 独立生成一个 GDST 文件
"""
import os, sys, re
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')


def xml_escape(s):
    if not s:
        return ''
    s = str(s)
    s = s.replace('&', '&amp;')
    s = s.replace('<', '&lt;')
    s = s.replace('>', '&gt;')
    s = s.replace('"', '&quot;')
    return s


def generate_gdst_xml(table_name, condition_headers, rows):
    """
    生成 GDST XML，支持动态条件列
    
    table_name: 表名
    condition_headers: 条件列名列表（不含结果列）
    rows: [{'row_num': int, 'description': str, 'conditions': [...], 'result': str}, ...]
    """
    parts = []
    parts.append('<decision-table52>\n')
    parts.append(f'      <tableName>{xml_escape(table_name)}</tableName>\n      \n      \n')

    parts.append('  <rowNumberCol>\n')
    parts.append('    <hideColumn>false</hideColumn>\n')
    parts.append('    <width>50</width>\n')
    parts.append('  </rowNumberCol>\n  \n\n      \n')

    parts.append('  <descriptionCol>\n')
    parts.append('    <hideColumn>false</hideColumn>\n')
    parts.append('    <width>200</width>\n')
    parts.append('  </descriptionCol>\n  \n\n')

    parts.append('      <metadataCols/>\n\n')
    parts.append('      <attributeCols/>\n\n')

    # 动态生成条件列
    parts.append('      <conditionPatterns>\n')
    parts.append('    <Pattern52>\n')
    parts.append('      <factType>CommonRuleFact</factType>\n')
    parts.append('      <boundName>f</boundName>\n')
    parts.append('      <isNegated>false</isNegated>\n')
    parts.append('      <conditions>\n')

    for i, header in enumerate(condition_headers):
        parts.append('  <condition-column52>\n    \n')
        if i == 0:
            # 第一列（规则名）特殊处理
            parts.append('  <typedDefaultValue>\n')
            parts.append('    <valueString></valueString>\n')
            parts.append('    <dataType>STRING</dataType>\n')
            parts.append('    <isOtherwise>false</isOtherwise>\n')
            parts.append('  </typedDefaultValue>\n')
            parts.append('    <hideColumn>false</hideColumn>\n')
            parts.append('    <width>213</width>\n')
            parts.append(f'    <header>{xml_escape(header)}</header>\n')
            parts.append('    <constraintValueType>1</constraintValueType>\n')
            parts.append('    <factField>rule</factField>\n')
            parts.append('    <fieldType>String</fieldType>\n')
            parts.append('    <operator>==</operator>\n')
        else:
            # 其他条件列
            parts.append('    <hideColumn>false</hideColumn>\n')
            parts.append('    <width>300</width>\n')
            parts.append(f'    <header>{xml_escape(header)}</header>\n')
            parts.append('    <constraintValueType>5</constraintValueType>\n')
            parts.append('    <factField></factField>\n')
            parts.append('    <fieldType></fieldType>\n')
            parts.append('    <operator></operator>\n')
        parts.append('    <parameters/>\n')
        parts.append('    <binding></binding>\n')
        parts.append('  </condition-column52>\n  \n')

    parts.append('  </conditions>\n')
    parts.append('      <window>\n')
    parts.append('        <parameters/>\n')
    parts.append('      </window>\n')
    parts.append('      <entryPointName></entryPointName>\n')
    parts.append('    </Pattern52>\n')
    parts.append('  </conditionPatterns>\n\n')

    # 动作列: 结果
    parts.append('      <actionCols>\n')
    parts.append('    <set-field-col52>\n')
    parts.append('      <typedDefaultValue>\n')
    parts.append('        <valueString></valueString>\n')
    parts.append('        <dataType>STRING</dataType>\n')
    parts.append('        <isOtherwise>false</isOtherwise>\n')
    parts.append('      </typedDefaultValue>\n')
    parts.append('      <hideColumn>false</hideColumn>\n')
    parts.append('      <width>250</width>\n')
    parts.append(f'      <header>结果</header>\n')
    parts.append('      <boundName>f</boundName>\n')
    parts.append('      <factField>s</factField>\n')
    parts.append('      <type>String</type>\n')
    parts.append('      <update>false</update>\n')
    parts.append('  </set-field-col52>\n')
    parts.append('  </actionCols>\n\n')

    # 审计日志
    parts.append('      <auditLog>\n')
    parts.append('        <filter class="org.drools.guvnor.client.modeldriven.dt52.auditlog.DecisionTableAuditLogFilter">\n')
    parts.append('          <acceptedTypes>\n')
    for t in ['INSERT_ROW', 'INSERT_COLUMN', 'DELETE_ROW', 'DELETE_COLUMN', 'UPDATE_COLUMN']:
        parts.append('            <entry>\n')
        parts.append(f'              <string>{t}</string>\n')
        parts.append('              <boolean>false</boolean>\n')
        parts.append('            </entry>\n')
    parts.append('          </acceptedTypes>\n')
    parts.append('        </filter>\n')
    parts.append('        <entries/>\n')
    parts.append('      </auditLog>\n\n')

    # 导入
    parts.append('      <imports><imports>\n')
    parts.append('  <org.kie.soup.project.datamodel.imports.Import>\n')
    parts.append('    <type>com.xiaomi.mifi.policy.model.CommonRuleFact</type>\n')
    parts.append('  </org.kie.soup.project.datamodel.imports.Import>\n')
    parts.append('  </imports></imports>\n\n')

    parts.append(f'      <packageName>com.xiaomi.mifi.{xml_escape(table_name)}</packageName>\n')
    parts.append('      <tableFormat>EXTENDED_ENTRY</tableFormat>\n')
    parts.append('      <hitPolicy>FIRST_HIT</hitPolicy>\n\n')

    # 数据行
    parts.append('      <data>')
    for row in rows:
        parts.append('<list>\n\n')

        # 行号
        parts.append('    <value>\n')
        parts.append(f'      <valueNumeric class="int">{row["row_num"]}</valueNumeric>\n')
        parts.append('      <valueString></valueString>\n')
        parts.append('      <dataType>NUMERIC_INTEGER</dataType>\n')
        parts.append('      <isOtherwise>false</isOtherwise>\n')
        parts.append('    </value>\n  \n')

        # 描述
        parts.append('    <value>\n')
        parts.append(f'      <valueString>{xml_escape(row["description"])}</valueString>\n')
        parts.append('      <dataType>STRING</dataType>\n')
        parts.append('      <isOtherwise>false</isOtherwise>\n')
        parts.append('    </value>\n  \n')

        # 动态条件列的值
        for cond_val in row.get('conditions', []):
            parts.append('    <value>\n')
            parts.append(f'      <valueString>{xml_escape(cond_val)}</valueString>\n')
            parts.append('      <dataType>STRING</dataType>\n')
            parts.append('      <isOtherwise>false</isOtherwise>\n')
            parts.append('    </value>\n  \n')

        # 结果
        parts.append('    <value>\n')
        parts.append(f'      <valueString>{xml_escape(row["result"])}</valueString>\n')
        parts.append('      <dataType>STRING</dataType>\n')
        parts.append('      <isOtherwise>false</isOtherwise>\n')
        parts.append('    </value>\n  \n')

        parts.append('  </list>')

    parts.append('</data>\n\n')
    parts.append('    </decision-table52>')

    return ''.join(parts)


def parse_excel_sheet(ws):
    """解析一个 sheet，返回 (table_name, condition_headers, data_rows)"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, [], []

    # 查找表头行（通常包含"规则名"和"结果"）
    header_row_idx = None
    for i, row in enumerate(rows[:10]):
        row_str = ' '.join([str(c) for c in row if c])
        if '规则名' in row_str and '结果' in row_str:
            header_row_idx = i
            break
    
    if header_row_idx is None:
        print(f"    警告: 找不到包含'规则名'和'结果'的表头行")
        return None, [], []

    headers = rows[header_row_idx]
    
    # 找到条件列（规则名到结果之间的列）
    rule_name_idx = None
    result_idx = None
    
    for i, h in enumerate(headers):
        h_str = str(h).strip() if h else ''
        if h_str == '规则名':
            rule_name_idx = i
        elif h_str == '结果':
            result_idx = i
    
    if rule_name_idx is None or result_idx is None:
        print(f"    警告: 找不到'规则名'或'结果'列")
        return None, [], []
    
    # 提取条件列头（规则名到结果列之间的列，不含结果列）
    condition_headers = []
    condition_indices = []
    for i in range(rule_name_idx, result_idx):  # 不包含result_idx
        h = headers[i]
        h_str = str(h).strip() if h else ''
        if h_str:
            condition_headers.append(h_str)
            condition_indices.append(i)
    
    # 获取描述（从元数据行或条件列中）
    description = ''
    if header_row_idx > 0:
        # 检查表头前一行是否包含描述
        meta_row = rows[header_row_idx - 1]
        for i, cell in enumerate(meta_row):
            if cell and str(cell).strip() == '描述':
                # 下一列是描述内容
                if i + 1 < len(meta_row) and meta_row[i + 1]:
                    description = str(meta_row[i + 1]).strip()
                break
    
    # 如果没有找到描述，尝试从条件列中提取
    if not description and len(condition_headers) > 0:
        # 通常描述在规则名列的值中，或者我们使用第一个条件列的表头
        description = condition_headers[0]
    
    print(f"    表头在第 {header_row_idx + 1} 行")
    print(f"    条件列: {condition_headers}")
    print(f"    描述: {description}")

    # 解析数据行
    data_rows = []
    row_num = 0

    for row in rows[header_row_idx + 1:]:
        # 跳过全空行
        if not any(cell for cell in row):
            continue

        # 提取条件值
        conditions = []
        for idx in condition_indices:
            if idx < len(row) and row[idx]:
                conditions.append(str(row[idx]).strip())
            else:
                conditions.append('')

        # 提取结果
        result = ''
        if result_idx < len(row) and row[result_idx]:
            result = str(row[result_idx]).strip()

        if not result:
            continue

        # 提取描述（从规则名列之前的列，通常是Col 1）
        # 找到规则名列之前的第一个非空列
        desc = ''
        for i in range(rule_name_idx):
            if i < len(row) and row[i]:
                desc = str(row[i]).strip()
                break

        # 处理结果格式（如果有权重信息，合并）
        # 假设结果格式已经是 code=weight 或需要转换
        if '=' not in result:
            # 如果结果没有=，添加默认权重
            result = f'{result}=60'

        row_num += 1
        data_rows.append({
            'row_num': row_num,
            'description': desc,
            'conditions': conditions,
            'result': result,
        })

    return condition_headers, data_rows


def process_excel_file(filepath):
    """处理一个 Excel 文件，返回 [(table_name, condition_headers, data_rows)]"""
    print(f"\n处理文件: {os.path.basename(filepath)}")

    wb = load_workbook(filepath, data_only=True)
    all_tables = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Sheet: {sheet_name}")

        result = parse_excel_sheet(ws)
        if result[0] is None:
            print(f"    跳过: 无法识别列结构")
            continue
        condition_headers, data_rows = result

        print(f"    数据行: {len(data_rows)}")

        if not data_rows:
            continue

        table_name = data_rows[0]['conditions'][0] if data_rows[0]['conditions'] else sheet_name
        # 清理表名中的特殊字符
        table_name = re.sub(r'[^\w]', '_', table_name).strip('_')
        
        all_tables.append((table_name, condition_headers, data_rows))

    wb.close()
    return all_tables


def main():
    # 确定输入路径
    if len(sys.argv) > 1:
        input_path = sys.argv[1]
    else:
        input_path = os.path.dirname(os.path.abspath(__file__))

    if not os.path.exists(input_path):
        print(f"路径不存在: {input_path}")
        return

    # 收集 Excel 文件
    if os.path.isfile(input_path):
        excel_files = [input_path]
    else:
        import glob
        excel_files = glob.glob(os.path.join(input_path, '*.xlsx'))
        excel_files += glob.glob(os.path.join(input_path, '*.xls'))

    if not excel_files:
        print(f"在 {input_path} 中没有找到 Excel 文件")
        return

    print(f"找到 {len(excel_files)} 个 Excel 文件")

    # 处理所有文件
    for f in excel_files:
        tables = process_excel_file(f)
        
        # 生成 GDST 文件
        print(f"\n{'='*60}")
        print(f"生成 GDST 文件")
        print(f"{'='*60}")

        base_dir = os.path.dirname(os.path.abspath(__file__))
        for table_name, condition_headers, data_rows in tables:
            # 重新编号
            for i, row in enumerate(data_rows):
                row['row_num'] = i + 1

            gdst_xml = generate_gdst_xml(table_name, condition_headers, data_rows)
            output_file = os.path.join(base_dir, f"{table_name}.gdst")

            with open(output_file, 'w', encoding='utf-8') as fw:
                fw.write(gdst_xml)

            print(f"  ✓ {table_name}.gdst ({len(data_rows)} 行)")
            print(f"    条件列: {condition_headers}")


if __name__ == '__main__':
    main()
